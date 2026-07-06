"""One-shot report: racing results so far for the gSheet breeze-up cohort.

Joins the ratings gSheet (the same sheet the daily email enriches from) to
Racing Post's sale catalogues on (sale, lot), then pulls each named horse's
form from the RP profile-tab endpoint and writes an xlsx with:

* ``Results`` — every sheet row plus Horse Name, Runs, First Time Out RPR
  and Peak RPR (RPR = Racing Post Rating, ``rpPostmark`` in RP's JSON).
* ``Breeze Rating Bands`` — per 10-point Breeze Rating band: lots, runners,
  average first-time-out RPR, average peak RPR and the highest RPR.
* ``Notes`` — column definitions and caveats.

Usage:

    .venv/bin/pip install -e ".[xlsx]"
    .venv/bin/python scripts/racing_results_xlsx.py --out "data/racing_results.xlsx"

Needs live Racing Post access (same bot-filter caveats as the daily email;
see README). Behind a TLS-intercepting proxy set ``RP_IMPERSONATE=chrome110``
— Chrome 124's post-quantum ClientHello is rejected by some MITM stacks —
and point ``CURL_CA_BUNDLE`` at the proxy CA.
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import time as wall
from dataclasses import dataclass
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from dailybreezeup import sheet as gsheet  # noqa: E402
from dailybreezeup.racing import rp_sales  # noqa: E402
from dailybreezeup.racing.rp_sales import _DOC_HEADERS, _XHR_HEADERS, BASE  # noqa: E402

log = logging.getLogger("racing_results_xlsx")

# The sheet's "Sale" column uses these short labels; map from the short names
# sheet.sale_short_name() derives from RP's verbose sale names.
_LABEL_OVERRIDES = {"Tattersalls Ireland": "Ireland"}


def make_session():
    """curl_cffi session; impersonation target and CA bundle overridable via env."""
    from curl_cffi import requests as cffi_requests

    impersonate = os.environ.get("RP_IMPERSONATE", "chrome124")
    verify = os.environ.get("CURL_CA_BUNDLE") or os.environ.get("REQUESTS_CA_BUNDLE")
    s = cffi_requests.Session(impersonate=impersonate, verify=verify or True)
    s.headers.update(_DOC_HEADERS)
    return s


def _get(session, url: str, referer: str, attempts: int = 4) -> str | None:
    for attempt in range(1, attempts + 1):
        try:
            r = session.get(url, headers=dict(_XHR_HEADERS, Referer=referer), timeout=30)
            if r.status_code == 200:
                return r.text
            log.warning("%s -> HTTP %s (attempt %d)", url, r.status_code, attempt)
        except Exception as exc:  # noqa: BLE001
            log.warning("%s -> %s (attempt %d)", url, exc, attempt)
        wall.sleep(2 * attempt)
    return None


@dataclass(frozen=True)
class HorseStats:
    runs: int
    fto_rpr: int | None
    peak_rpr: int | None


def fetch_form_stats(session, uid: int, *, sleep: float) -> HorseStats | None:
    """First-time-out and peak RPR from the profile form tab (all runs to date)."""
    url = f"{BASE}/profile/tab/horse/{uid}/x/form"
    body = _get(session, url, f"{BASE}/profile/horse/{uid}")
    wall.sleep(sleep)
    if body is None:
        return None
    try:
        runs = json.loads(body).get("form") or []
    except json.JSONDecodeError:
        log.warning("form tab for %s returned non-JSON", uid)
        return None
    runs = sorted(runs, key=lambda r: r.get("raceDatetime") or "")
    rprs = [r["rpPostmark"] for r in runs if r.get("rpPostmark") is not None]
    fto = runs[0].get("rpPostmark") if runs else None
    return HorseStats(runs=len(runs), fto_rpr=fto, peak_rpr=max(rprs) if rprs else None)


def band_of(rating: float) -> str:
    if rating >= 100:
        return "100+"
    if rating < 40:
        return "<40"
    lo = int(rating // 10) * 10
    return f"{lo}-{lo + 9}"


BAND_ORDER = ["100+", "90-99", "80-89", "70-79", "60-69", "50-59", "40-49", "<40"]

_NOTES = [
    "Racing results for the breeze-up sale cohort, joined to Racing Post form records.",
    "",
    "Results sheet:",
    "  Horse Name — racing name per Racing Post's sale catalogue"
    " (blank = still unnamed/unregistered).",
    "  Runs — completed runs to date. Blank = no RP horse record yet; 0 = named but unraced.",
    "  First Time Out RPR — Racing Post Rating achieved on debut."
    " Blank if unraced or debut not yet rated.",
    "  Peak RPR — highest RPR achieved in any start so far.",
    "",
    "Breeze Rating Bands sheet:",
    "  Lots — sheet rows in the band with a Breeze Rating.",
    "  Runners — of those, horses that have raced at least once.",
    "  Averages are over horses with a rated run;"
    " Highest RPR is the best single figure in the band.",
    "",
    "Notes: a very recent run may not have an RPR published yet. Horses re-offered at a",
    "second sale appear once per sale row (each breeze carries its own rating).",
]


def build_workbook(rows, lot_map, stats_by_uid, out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F4E79")

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    header = [
        "Year", "Sale", "Lot", "Sex", "Sire", "Dam", "Vendor", "Buyer", "Price",
        "OT Diff/M", "OT Rank", "SL 1F", "SL GO", "Breeze Rating", "Precocity Rating",
        "Horse Name", "Runs", "First Time Out RPR", "Peak RPR",
    ]
    ws.append(header)
    for cell in ws[1]:
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = border

    def price_num(price: str) -> int | str | None:
        s = price.replace(",", "").strip()
        return int(s) if s.isdigit() else (price or None)

    per_band = {b: {"lots": 0, "raced": 0, "ftos": [], "peaks": []} for b in BAND_ORDER}
    for row in rows:
        lot = lot_map.get((row.sale, str(row.lot)))
        stats = stats_by_uid.get(lot.horse_uid) if lot and lot.horse_uid else None
        name = lot.horse_name if lot else ""
        ws.append([
            row.year, row.sale, row.lot, row.sex, row.sire, row.dam, row.vendor, row.buyer,
            price_num(row.price), row.ot_diff_m, row.ot_rank, row.sl_1f, row.sl_go,
            row.breeze_rating, row.precocity_rating,
            name,
            stats.runs if stats else (0 if name else None),
            stats.fto_rpr if stats else None,
            stats.peak_rpr if stats else None,
        ])
        if row.breeze_rating is not None:
            d = per_band[band_of(row.breeze_rating)]
            d["lots"] += 1
            if stats and stats.runs:
                d["raced"] += 1
            if stats and stats.fto_rpr is not None:
                d["ftos"].append(stats.fto_rpr)
            if stats and stats.peak_rpr is not None:
                d["peaks"].append(stats.peak_rpr)

    n_cols = len(header)
    for r in range(2, ws.max_row + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = border
        price_cell = ws.cell(row=r, column=header.index("Price") + 1)
        if isinstance(price_cell.value, int):
            price_cell.number_format = "#,##0"
    widths = [6, 9, 6, 5, 22, 22, 34, 40, 10, 10, 9, 7, 8, 13, 15, 22, 6, 17, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{ws.max_row}"

    def avg(xs: list[int]) -> float | None:
        return round(sum(xs) / len(xs), 1) if xs else None

    ws2 = wb.create_sheet("Breeze Rating Bands")
    ws2.append(["Breeze Rating Band", "Lots", "Runners",
                "Avg First Time Out RPR", "Avg Peak RPR", "Highest RPR"])
    for cell in ws2[1]:
        cell.font = head_font
        cell.fill = head_fill
        cell.border = border
    all_ftos: list[int] = []
    all_peaks: list[int] = []
    total_lots = total_raced = 0
    for b in BAND_ORDER:
        d = per_band[b]
        ws2.append([b, d["lots"], d["raced"], avg(d["ftos"]), avg(d["peaks"]),
                    max(d["peaks"]) if d["peaks"] else None])
        all_ftos += d["ftos"]
        all_peaks += d["peaks"]
        total_lots += d["lots"]
        total_raced += d["raced"]
    ws2.append(["All rated lots", total_lots, total_raced, avg(all_ftos), avg(all_peaks),
                max(all_peaks) if all_peaks else None])
    for r in range(2, ws2.max_row + 1):
        for c in range(1, 7):
            cell = ws2.cell(row=r, column=c)
            cell.border = border
            if r == ws2.max_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="DDEBF7")
    for i, w in enumerate([18, 8, 10, 21, 13, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Notes")
    for line in _NOTES:
        ws3.append([line])
    ws3.column_dimensions["A"].width = 110

    wb.save(out_path)


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--year", type=int, default=2026)
    ap.add_argument("--out", type=Path, default=Path("data/racing_results.xlsx"))
    ap.add_argument("--sheet-url", default=os.environ.get("SHEET_CSV_URL")
                    or gsheet.DEFAULT_SHEET_CSV_URL)
    ap.add_argument("--sleep", type=float, default=0.45,
                    help="pause between RP requests (politeness)")
    args = ap.parse_args()

    rows = gsheet.fetch_sheet(args.sheet_url)
    log.info("sheet: %d rows", len(rows))

    session = make_session()
    lot_map: dict[tuple[str, str], rp_sales.SaleLot] = {}
    for sale in rp_sales.discover_sales(args.year, session=session):
        short = gsheet.sale_short_name(sale.sale_name)
        if short is None:
            log.info("skipping %s (no sheet label)", sale.sale_name)
            continue
        label = _LABEL_OVERRIDES.get(short, short)
        lots = rp_sales.fetch_lots(sale, session=session, sleep=args.sleep)
        log.info("%s (%s): %d lots", sale.sale_name, label, len(lots))
        for lot in lots:
            lot_map[(label, lot.display_lot)] = lot

    uids = sorted({lot.horse_uid for lot in lot_map.values() if lot.horse_uid})
    log.info("fetching form for %d named horses", len(uids))
    stats_by_uid: dict[int, HorseStats] = {}
    for i, uid in enumerate(uids, 1):
        stats = fetch_form_stats(session, uid, sleep=args.sleep)
        if stats is not None:
            stats_by_uid[uid] = stats
        if i % 50 == 0:
            log.info("  %d/%d", i, len(uids))

    args.out.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(rows, lot_map, stats_by_uid, args.out)
    raced = sum(1 for s in stats_by_uid.values() if s.runs)
    log.info("saved %s (%d named, %d raced)", args.out, len(uids), raced)


if __name__ == "__main__":
    main()
